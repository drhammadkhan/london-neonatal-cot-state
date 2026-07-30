#!/usr/bin/env python3
"""
Regenerate "London Neonatal Cot State.html" from a Cot State Excel export.

Usage:
    python3 generate.py                       # auto-picks newest "Cot State*.xlsx" here
    python3 generate.py "Cot State - 05.08.2026.xlsx"
    python3 generate.py path/to/file.xlsx  path/to/output.html

The output is a single, self-contained HTML file (map tiles + fonts load online).
Only the "Cotstate" sheet layout is assumed:
  col A  full name        col S  RAG status (col 20)
  col C  short name       col V  impending admissions (22)
  col D  level            col W  planned discharges (23)
  col E  status           col Y  comment (25)
  col G/H IC/HD total/avail        col Q  overall "avail / total" (17)
  col L/M SC funded/avail          col R  overall pct (18)
Network subtotal rows carry the network name in col C and an empty col A.
"""
import sys, os, glob, json, datetime, re
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, 'template.html')
DEFAULT_OUT = os.path.join(HERE, 'London Neonatal Cot State.html')

# --- Unit map locations. Single source of truth = geo.json (shared with the Builder).
#     Add new units there, not here.
GEO_PATH = os.path.join(HERE, 'geo.json')


def load_geo():
    if not os.path.exists(GEO_PATH):
        sys.exit('geo.json is missing — it must sit next to generate.py.')
    with open(GEO_PATH, encoding='utf-8') as f:
        raw = json.load(f)
    return {k: tuple(v) for k, v in raw.items() if not k.startswith('_')}


GEO = load_geo()
LONDON_CENTRE = (51.5074, -0.1278)
VALID_LEVELS = {'NICU', 'LNU', 'SCBU'}

# Network name (as it appears in the subtotal row) -> canonical name + abbreviation
NET_CANON = {
 'north central london': ('North Central London', 'NCL'),
 'north east london':    ('North East London', 'NEL'),
 'north west london':    ('North West London', 'NWL'),
 'north west':           ('North West London', 'NWL'),
 'south east london':    ('South East London', 'SEL'),
 'south west london':    ('South West London', 'SWL'),
}


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def parse_pair(v):
    """Parse 'avail / total' -> (avail, total). Returns (None, None) if unavailable."""
    if v is None:
        return (None, None)
    m = re.findall(r'-?\d+', str(v))
    if len(m) >= 2:
        return (int(m[0]), int(m[1]))
    return (None, None)


def to_int(v):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def to_float(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ('--', '-- / --') else None


def find_report_datetime(ws):
    """Find the report timestamp (a datetime in the header area)."""
    for r in range(1, 6):
        for c in range(1, 12):
            v = cell(ws, r, c)
            if isinstance(v, datetime.datetime):
                return v
    return None


def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Cotstate']

    units, networks = [], []
    buffer = []          # units accumulated since the last subtotal
    warnings = []
    london = None

    for r in range(1, ws.max_row + 1):
        full = clean(cell(ws, r, 1))
        short = clean(cell(ws, r, 3))
        if not short:
            continue

        key = short.lower().strip()

        # London grand-total row
        if key == 'london':
            av, fn = parse_pair(cell(ws, r, 17))
            ic_av, ic_tot = parse_pair(cell(ws, r, 9))
            sc_av, sc_tot = parse_pair(cell(ws, r, 14))
            pct = to_float(cell(ws, r, 18))
            london = {
                'funded': fn, 'available': av,
                'occupied': (fn - av) if (fn is not None and av is not None) else None,
                'pct': round(pct * 100) if pct is not None else None,
                'ichdAvail': ic_av, 'ichdTotal': ic_tot, 'scAvail': sc_av, 'scTotal': sc_tot,
            }
            continue

        # Network subtotal row: name in col C, no full name in col A
        if key in NET_CANON and not full:
            name, abbr = NET_CANON[key]
            av, fn = parse_pair(cell(ws, r, 17))
            pct = to_float(cell(ws, r, 18))
            for u in buffer:
                u['net'] = name
                u['netAbbr'] = abbr
            networks.append({
                'name': name, 'abbr': abbr, 'funded': fn, 'available': av,
                'pct': pct,
                'ichdAvail': parse_pair(cell(ws, r, 9))[0], 'ichdTotal': parse_pair(cell(ws, r, 9))[1],
                'scAvail': parse_pair(cell(ws, r, 14))[0], 'scFunded': parse_pair(cell(ws, r, 14))[1],
            })
            buffer = []
            continue

        # Otherwise: a unit row (needs a full name and a real neonatal level)
        level = clean(cell(ws, r, 4))
        if not full or (level or '').upper() not in VALID_LEVELS:
            continue

        av, fn = parse_pair(cell(ws, r, 17))
        ic_av, ic_tot = parse_pair(cell(ws, r, 9))
        sc_av, sc_tot = parse_pair(cell(ws, r, 14))
        lat_lng = GEO.get(short)
        if lat_lng is None:
            warnings.append(short)
            lat_lng = LONDON_CENTRE

        u = {
            'full': full, 'short': short, 'level': level,
            'status': clean(cell(ws, r, 5)), 'net': None, 'netAbbr': None,
            'lat': lat_lng[0], 'lng': lat_lng[1],
            'ichdTotal': ic_tot, 'ichdAvail': ic_av,
            'scFunded': sc_tot, 'scAvail': sc_av,
            'overall': clean(cell(ws, r, 17)), 'overallPct': to_float(cell(ws, r, 18)) or 0,
            'rag': clean(cell(ws, r, 20)),
            'admissions': to_int(cell(ws, r, 22)), 'discharges': to_int(cell(ws, r, 23)),
            'comment': clean(cell(ws, r, 25)),
            # funded (capacity) is reliable even when a unit doesn't report; fall back to it.
            'funded': fn if fn is not None else to_int(cell(ws, r, 27)),
            # available: ONLY from the reported overall. If the unit didn't submit
            # ('-- / --'), leave it None (→ treated as stale) rather than trusting the
            # "Total Available" column, which then just mirrors funded (looks fully empty).
            'available': av,
        }
        units.append(u)
        buffer.append(u)

    # Report date/time
    dt = find_report_datetime(ws)
    if dt:
        # strip leading zero on day for a natural "29 July 2026"
        date_str = dt.strftime('%A ') + str(dt.day) + dt.strftime(' %B %Y')
        time_str = dt.strftime('%H:%M')
    else:
        date_str, time_str = 'Latest report', ''

    data = {
        'meta': {
            'title': 'London Neonatal Network',
            'date': date_str, 'time': time_str,
            'note': 'Anonymised data — from daily Badgernet Cot Bureau submissions.',
            'source': 'London Neonatal Network · londonneonatalnetwork.org.uk',
        },
        'london': london or {},
        'networks': networks,
        'units': units,
    }
    return data, warnings


# Fields that change day to day (stored per frame) vs stable identity/location.
DYNAMIC = ('status', 'ichdTotal', 'ichdAvail', 'scFunded', 'scAvail', 'overall',
           'overallPct', 'rag', 'admissions', 'discharges', 'comment', 'funded', 'available')
STATIC = ('full', 'short', 'level', 'net', 'netAbbr', 'lat', 'lng')


def date_key(path):
    """Sort key by the report DATE in the filename (DD.MM.YYYY); mtime as fallback."""
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', os.path.basename(path))
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return (1, datetime.date(y, mo, d).toordinal())
        except ValueError:
            pass
    return (0, os.path.getmtime(path))


def build_payload(frames):
    """frames: list of extract() dicts, oldest→newest. Latest is the map's default day."""
    latest = frames[-1]
    static = [{k: u.get(k) for k in STATIC} for u in latest['units']]
    frame_out = []
    for data in frames:
        um = {u['short']: {k: u.get(k) for k in DYNAMIC} for u in data['units']}
        frame_out.append({'date': data['meta']['date'], 'time': data['meta']['time'],
                          'london': data['london'], 'units': um})
    return {'meta': latest['meta'], 'networks': latest['networks'],
            'london': latest['london'], 'units': static, 'frames': frame_out}


def main():
    args = sys.argv[1:]
    out = args[1] if len(args) >= 2 else DEFAULT_OUT

    if len(args) >= 1:
        files = [args[0]]                      # explicit single file → single day
    else:
        files = [c for c in glob.glob(os.path.join(HERE, 'Cot State*.xlsx'))
                 if not os.path.basename(c).startswith('~$')]
        if not files:
            sys.exit('No "Cot State*.xlsx" found here. Pass the file path as an argument.')
        files.sort(key=date_key)               # oldest → newest

    if not os.path.exists(TEMPLATE):
        sys.exit('template.html is missing — it must sit next to generate.py.')

    frames, warnings = [], []
    for f in files:
        print(f'Reading:   {os.path.basename(f)}')
        data, warns = extract(f)
        frames.append(data)
        for w in warns:
            if w not in warnings:
                warnings.append(w)

    payload_obj = build_payload(frames)
    template = open(TEMPLATE, encoding='utf-8').read()
    payload = json.dumps(payload_obj, indent=1, ensure_ascii=False)
    html = template.replace('/*DATA*/', payload, 1)
    with open(out, 'w', encoding='utf-8') as f:
        f.write(html)

    latest = frames[-1]
    def short_date(d):
        p = d['meta']['date'].split()      # e.g. ['Thursday','24','July','2026']
        return f'{p[1]} {p[2][:3]}' if len(p) >= 3 else d['meta']['date']
    print(f'Days:      {len(frames)}  ({" · ".join(short_date(d) for d in frames)})')
    print(f'Units:     {len(latest["units"])}   Networks: {len(latest["networks"])}')
    print(f'Latest:    {latest["meta"]["date"]} {latest["meta"]["time"]}')
    if latest['london']:
        L = latest['london']
        print(f'London:    {L.get("available")}/{L.get("funded")} available ({L.get("pct")}%)')
    if warnings:
        print('\n  WARNING — no map location for these units (placed at London centre):')
        for w in warnings:
            print(f'    • {w}   → add coordinates to geo.json')
    print(f'\nWrote:     {out}')


if __name__ == '__main__':
    main()
